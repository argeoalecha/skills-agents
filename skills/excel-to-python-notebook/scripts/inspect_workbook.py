"""Extract everything needed to plan an Excel to Python conversion.

Two passes over the workbook: formulas (data_only=False) and cached values
(data_only=True). Reports the facts that decide conversion strategy — which cells
carry logic, which functions appear, what is volatile, what has no cached value to
validate against, whether VBA or iterative calculation is involved.

Usage:
    python3 inspect_workbook.py <workbook> [--sheet NAME] [--json OUT] [--formulas N]

Exit codes: 0 ok, 2 workbook unreadable.
"""

import argparse
import json
import re
import sys
import warnings
import zipfile
from collections import Counter
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# Recalculated on every edit, so a cached value is one frozen draw, not a target.
VOLATILE = {"RAND", "RANDBETWEEN", "RANDARRAY", "NOW", "TODAY", "OFFSET", "INDIRECT", "INFO", "CELL"}

# Excel functions with no direct numpy/scipy equivalent worth flagging early.
NEEDS_CARE = {
    "GAMMALN", "NORMSINV", "NORMSDIST", "NORMINV", "NORMDIST", "CHIINV", "CHIDIST",
    "CHISQ.INV", "CHISQ.INV.RT", "CHISQ.DIST", "TINV", "FINV", "WEIBULL", "EXPON.DIST",
    "LINEST", "TREND", "SLOPE", "INTERCEPT", "RSQ", "FORECAST", "VLOOKUP", "HLOOKUP",
    "INDEX", "MATCH", "SUMPRODUCT", "FREQUENCY", "PERCENTILE", "QUARTILE",
}

FUNC_RE = re.compile(r"\b([A-Z][A-Z0-9_.]*)\s*\(")

# A broken source formula must never be silently reproduced in Python.
ERROR_VALUES = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


def formula_text(value):
    """openpyxl yields str for normal formulas, ArrayFormula objects for CSE formulas."""
    if isinstance(value, str):
        return value if value.startswith("=") else None
    text = getattr(value, "text", None)
    return text if isinstance(text, str) and text.startswith("=") else None


def grid_sheets(wb):
    """Chartsheets have no cells; skip them without crashing on .iter_rows."""
    return [(name, wb[name]) for name in wb.sheetnames if hasattr(wb[name], "iter_rows")]


def inspect(path, only_sheet=None, formula_samples=12):
    path = Path(path)
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        sys.exit(f"cannot read {path}: {exc}")

    calc = getattr(wb_f, "calculation", None)
    report = {
        "workbook": str(path),
        "sheets_all": wb_f.sheetnames,
        "chartsheets": [n for n in wb_f.sheetnames if not hasattr(wb_f[n], "iter_rows")],
        "has_vba": False,
        "calc_mode": getattr(calc, "calcMode", None),
        "iterative_calc": bool(getattr(calc, "iterate", False)),
        "iterate_count": getattr(calc, "iterateCount", None),
        "defined_names": [],
        "external_links": 0,
        "sheets": [],
    }

    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            report["has_vba"] = "xl/vbaProject.bin" in names
    except Exception:
        pass

    try:
        report["defined_names"] = sorted(wb_f.defined_names.keys())
    except Exception:
        pass
    try:
        report["external_links"] = len(wb_f._external_links or [])
    except Exception:
        pass

    for name, ws in grid_sheets(wb_f):
        if only_sheet and name != only_sheet:
            continue
        ws_v = wb_v[name] if hasattr(wb_v[name], "iter_rows") else None

        funcs = Counter()
        n_formula = n_const = n_uncached = 0
        volatile_cells = []
        uncached_cells = []
        error_cells = []
        samples = []
        unique_shapes = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cached_val = ws_v[cell.coordinate].value if ws_v is not None else None
                if isinstance(cached_val, str) and cached_val.startswith(ERROR_VALUES):
                    if len(error_cells) < 25:
                        error_cells.append(f"{cell.coordinate}={cached_val}")

                ftext = formula_text(cell.value)
                if ftext is None:
                    n_const += 1
                    continue
                n_formula += 1
                found = set(FUNC_RE.findall(ftext))
                funcs.update(found)

                if any(e in ftext for e in ERROR_VALUES) and len(error_cells) < 25:
                    error_cells.append(f"{cell.coordinate} formula:{ftext[:40]}")

                if found & VOLATILE:
                    volatile_cells.append(cell.coordinate)

                if ws_v is not None and cached_val is None:
                    n_uncached += 1
                    if len(uncached_cells) < 25:
                        uncached_cells.append(cell.coordinate)

                # Collapse row-repeated formulas: strip digits to find the shape.
                shape = re.sub(r"\d+", "#", ftext)
                if shape not in unique_shapes:
                    unique_shapes[shape] = cell.coordinate
                    if len(samples) < formula_samples:
                        samples.append({
                            "cell": cell.coordinate,
                            "formula": ftext,
                            "cached": cached_val,
                        })

        report["sheets"].append({
            "name": name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "n_formula_cells": n_formula,
            "n_constant_cells": n_const,
            "n_unique_formula_shapes": len(unique_shapes),
            "n_uncached_formula_cells": n_uncached,
            "uncached_sample": uncached_cells,
            "n_volatile_cells": len(volatile_cells),
            "volatile_sample": volatile_cells[:10],
            "error_cells": error_cells,
            "functions": dict(funcs.most_common()),
            "functions_needing_care": sorted(set(funcs) & NEEDS_CARE),
            "distinct_formulas": samples,
        })

    return report


def render(report):
    out = []
    add = out.append
    add(f"WORKBOOK  {report['workbook']}")
    add(f"  sheets: {len(report['sheets_all'])}  (chartsheets skipped: {report['chartsheets'] or 'none'})")
    add(f"  VBA project: {'YES — macro logic lives outside cells' if report['has_vba'] else 'no'}")
    add(f"  calc mode: {report['calc_mode']}   iterative: {report['iterative_calc']} ({report['iterate_count']})")
    if report["external_links"]:
        add(f"  external links: {report['external_links']}  — resolve before converting")
    if report["defined_names"]:
        add(f"  defined names: {', '.join(report['defined_names'][:12])}")
    add("")

    for s in report["sheets"]:
        add(f"SHEET  {s['name']}   [{s['dimensions']}]")
        add(f"  formula cells {s['n_formula_cells']}  (unique shapes {s['n_unique_formula_shapes']})  constants {s['n_constant_cells']}")
        if s["n_uncached_formula_cells"]:
            add(f"  UNCACHED formula cells: {s['n_uncached_formula_cells']} — not validatable by parity")
            add(f"    e.g. {', '.join(s['uncached_sample'][:8])}")
        if s["n_volatile_cells"]:
            add(f"  VOLATILE cells: {s['n_volatile_cells']} — cached values are one frozen draw, use tolerance checks")
            add(f"    e.g. {', '.join(s['volatile_sample'][:8])}")
        if s["error_cells"]:
            add(f"  ERROR VALUES in source: {len(s['error_cells'])} — the workbook itself is broken here, do not reproduce silently")
            add(f"    e.g. {', '.join(s['error_cells'][:5])}")
        if s["functions_needing_care"]:
            add(f"  functions needing a deliberate Python mapping: {', '.join(s['functions_needing_care'])}")
        if s["functions"]:
            top = ", ".join(f"{k}x{v}" for k, v in list(s["functions"].items())[:12])
            add(f"  function use: {top}")
        if s["distinct_formulas"]:
            add("  distinct formulas:")
            for d in s["distinct_formulas"]:
                cached = d["cached"]
                shown = f"  -> {cached!r}" if cached is not None else "  -> (no cached value)"
                add(f"    {d['cell']:>6}  {d['formula']}{shown}")
        add("")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--sheet", help="inspect only this sheet")
    ap.add_argument("--json", help="write full report as JSON to this path")
    ap.add_argument("--formulas", type=int, default=12, help="distinct formulas to show per sheet")
    args = ap.parse_args()

    report = inspect(args.workbook, args.sheet, args.formulas)
    print(render(report))
    if args.json:
        Path(args.json).write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        print(f"[json written to {args.json}]")


if __name__ == "__main__":
    main()
