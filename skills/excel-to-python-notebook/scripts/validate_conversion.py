"""Three-tier validation of a Python conversion against its source workbook.

Import this into a converted notebook. It reads cached values out of the original
workbook and records one of three verdicts per check:

    exact()         deterministic cell — Python must match Excel to rtol
    tolerance()     volatile/simulated cell — Excel's cached value is one frozen
                    draw and its RNG cannot be seed-matched, so only agreement
                    within a stated tolerance is meaningful
    unvalidatable() no cached value exists (VBA-written range, or the source cell
                    is broken) — recorded as a known gap, never silently skipped

Usage in a notebook:

    from validate_conversion import Validator
    v = Validator("Session 2.4.2 - MLE.xlsx")
    v.exact("beta (MLE)", beta_py, v.cell("CL-Weibull", "D2"))
    v.tolerance("P(system up)", p_py, v.cell("SERIES", "G5"), rtol=0.05)
    v.unvalidatable("SERIES!C6:C12", "VBA-written at runtime, no cached value")
    v.report()

CLI (quick look at what a sheet cached):

    python3 validate_conversion.py <workbook> --sheet NAME --range B15:D20
"""

import argparse
import math
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

ERROR_VALUES = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")

PASS, FAIL, GAP = "PASS", "FAIL", "GAP"

# Excel's legacy iterative solvers for inverse distributions are accurate to roughly
# 1e-7 relative, not machine precision. Measured against scipy on real workbooks:
# CHIINV drifts up to ~3e-8, while NORMSINV agrees to ~8e-16. When a check fails only
# at the 1e-8 level on one of these, scipy is the more accurate side — loosen rtol to
# RTOL_LEGACY_STAT and say so in the notebook rather than bending the Python to match.
RTOL_LEGACY_STAT = 1e-6
LEGACY_STAT_FUNCS = {
    "CHIINV", "CHIDIST", "CHISQ.INV", "CHISQ.INV.RT", "TINV", "FINV",
    "GAMMAINV", "BETAINV", "BINOMDIST", "POISSON",
}


class Validator:
    """Collects verdicts comparing Python results against a workbook's cached values."""

    def __init__(self, workbook, rtol=1e-9):
        self.path = Path(workbook)
        self.default_rtol = rtol
        self._wb = openpyxl.load_workbook(self.path, data_only=True)
        self.checks = []

    # -- reading the source -------------------------------------------------

    def cell(self, sheet, coord):
        """Cached value of one cell. Returns None when Excel never cached one."""
        value = self._wb[sheet][coord].value
        if isinstance(value, str) and value.startswith(ERROR_VALUES):
            return None
        return value

    def range(self, sheet, ref, drop_none=True):
        """Cached values of a range as a flat list, e.g. range('Data', 'B15:B1014')."""
        out = []
        for row in self._wb[sheet][ref]:
            cells = row if isinstance(row, tuple) else (row,)
            for c in cells:
                v = c.value
                if isinstance(v, str) and v.startswith(ERROR_VALUES):
                    v = None
                if v is None and drop_none:
                    continue
                out.append(v)
        return out

    # -- recording verdicts -------------------------------------------------

    def legacy_stat(self, label, python_value, excel_value):
        """Deterministic cell computed by a low-precision legacy Excel solver.

        Use for CHIINV/TINV/FINV/GAMMAINV and friends — see RTOL_LEGACY_STAT.
        """
        return self.exact(label, python_value, excel_value, rtol=RTOL_LEGACY_STAT)

    def exact(self, label, python_value, excel_value, rtol=None):
        """Deterministic cell: Python must reproduce Excel to floating-point tolerance."""
        rtol = self.default_rtol if rtol is None else rtol
        if excel_value is None:
            return self.unvalidatable(label, "no cached value in source workbook")
        ok = self._close(python_value, excel_value, rtol)
        self.checks.append((PASS if ok else FAIL, label, python_value, excel_value,
                            f"exact rtol={rtol:g}"))
        return ok

    def tolerance(self, label, python_value, excel_value, rtol=0.05):
        """Volatile/simulated cell: agreement within rtol is the strongest claim available."""
        if excel_value is None:
            return self.unvalidatable(label, "no cached value in source workbook")
        ok = self._close(python_value, excel_value, rtol)
        self.checks.append((PASS if ok else FAIL, label, python_value, excel_value,
                            f"tolerance rtol={rtol:g} (Excel RNG not seed-matchable)"))
        return ok

    def unvalidatable(self, label, reason):
        """Record a known validation gap so it appears in the report instead of vanishing."""
        self.checks.append((GAP, label, None, None, reason))
        return None

    @staticmethod
    def _close(a, b, rtol):
        try:
            a, b = float(a), float(b)
        except (TypeError, ValueError):
            return a == b
        if math.isnan(a) or math.isnan(b):
            return False
        return math.isclose(a, b, rel_tol=rtol, abs_tol=1e-12)

    # -- output -------------------------------------------------------------

    def report(self, verbose=True):
        """Print the verdict table. Returns True when nothing FAILed."""
        n_pass = sum(1 for c in self.checks if c[0] == PASS)
        n_fail = sum(1 for c in self.checks if c[0] == FAIL)
        n_gap = sum(1 for c in self.checks if c[0] == GAP)

        if verbose:
            print(f"Validation against {self.path.name}")
            print("-" * 78)
            for verdict, label, py, xl, note in self.checks:
                if verdict == GAP:
                    print(f"  {verdict:4}  {label}")
                    print(f"        not validatable: {note}")
                else:
                    print(f"  {verdict:4}  {label}")
                    print(f"        python={py!r}  excel={xl!r}  [{note}]")
            print("-" * 78)
            print(f"  {n_pass} passed, {n_fail} failed, {n_gap} unvalidatable")
            if n_gap and not n_fail:
                print("  Note: unvalidatable entries are known gaps, not silent passes.")

        return n_fail == 0


def main():
    ap = argparse.ArgumentParser(description="Inspect cached values in a workbook range.")
    ap.add_argument("workbook")
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--range", dest="ref", required=True, help="e.g. B15:D20")
    args = ap.parse_args()

    v = Validator(args.workbook)
    values = v.range(args.sheet, args.ref, drop_none=False)
    print(f"{args.sheet}!{args.ref} — {len(values)} cells, "
          f"{sum(1 for x in values if x is None)} with no cached value")
    for val in values:
        print(f"  {val!r}")


if __name__ == "__main__":
    main()
